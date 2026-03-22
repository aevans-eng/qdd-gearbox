"""Build the QDD Requirements Traceability Matrix spreadsheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Colors
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
INPROG_FILL = PatternFill("solid", fgColor="FFEB9C")
GAP_FILL = PatternFill("solid", fgColor="F4B084")
PARTIAL_FILL = PatternFill("solid", fgColor="FFD966")
LIGHT_BLUE = PatternFill("solid", fgColor="D6E4F0")
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
PRIMARY_FILL = PatternFill("solid", fgColor="BDD7EE")
SUPPORT_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = openpyxl.Workbook()

# ── Helper ──────────────────────────────────────────────────────────────
def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"

def auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        lengths = []
        for cell in col:
            if cell.value:
                for line in str(cell.value).split("\n"):
                    lengths.append(len(line))
        best = min(max(max(lengths, default=min_w), min_w), max_w)
        ws.column_dimensions[letter].width = best + 2

def add_validation(ws, col_letter, options, row_start=2, row_end=50):
    dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
    dv.error = "Pick from the list"
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{row_start}:{col_letter}{row_end}")


# ═══════════════════════════════════════════════════════════════════════
# SHEET 1: Requirements
# ═══════════════════════════════════════════════════════════════════════
ws_req = wb.active
ws_req.title = "Requirements"
req_headers = [
    "Req ID", "Requirement", "Description", "Target Value",
    "Type", "Priority Rank", "Verification Method",
    "Status", "Verified By", "Notes",
]
ws_req.append(req_headers)
style_header(ws_req, len(req_headers))

requirements = [
    ("R-01", "Backlash", "Angular backlash at output shaft", "≤ 0.5°",
     "Hard", 1, "Test", "Not Verified", "T-012",
     "Primary QDD metric. Measured with dial indicator."),
    ("R-02", "Cost", "Total BOM cost under budget", "< $120 CAD",
     "Hard", 2, "Inspection", "Not Verified", "BOM inspection",
     "Sum of purchased + printed parts cost"),
    ("R-03", "DFM/DFA", "All off-the-shelf + 3D printed parts; assemblable without complex tools",
     "Off-the-shelf + FDM, no special tools",
     "Hard", 3, "Inspection", "Not Verified", "Design review",
     "Assembly with basic hex keys only"),
    ("R-04", "Backdrivability", "Friction torque with motor disconnected",
     "< 1 Nm friction torque",
     "Hard", 4, "Test", "Not Verified", "T-013, T-008, T-006",
     "Changed from vague 'backdriveable under human impulse' to measurable spec (Mar 11). Industry standard for QDD."),
    ("R-05", "Durability", "Should not break itself in under 1 minute",
     "No self-destruction in < 1 min",
     "Hard", 5, "Test", "Not Verified", "",
     "Phase 5 only. Combination of software + hardware limits."),
    ("R-06", "Peak Torque", "Peak torque output at gearbox output shaft",
     "≥ 16 Nm",
     "Hard", 6, "Test", "Not Verified", "",
     "Phase 5 only. Needs load cell or heavy weights."),
    ("R-07", "Continuous Torque", "Sustained torque output capacity",
     "≥ 12 Nm",
     "Hard", 7, "Test", "Not Verified", "",
     "Phase 5 only. Needs thermal steady-state test."),
    ("R-08", "Thermal Performance", "Housing should not overheat under sustained use",
     "No melting in < 5 min",
     "Hard", 8, "Test", "Not Verified", "",
     "Phase 5 only. PLA Tg ~60°C. Health monitoring tracks informally."),
    ("R-09", "Efficiency", "Power transmission efficiency through gearbox",
     "> 90%",
     "Soft", 9, "Test", "Not Verified", "T-006 (related)",
     "Phase 5 only. T-006 measures friction but not output power directly."),
    ("R-10", "Weight", "Total assembly mass",
     "< 2 kg",
     "Soft", 10, "Inspection", "Not Verified", "",
     "Trivial — kitchen scale measurement. Add as 1-min check in Phase 1."),
    ("R-11", "Speed", "Maximum continuous output RPM",
     "≥ 600 RPM continuous",
     "Soft", 11, "Test", "Not Verified", "T-016 (related)",
     "T-016 measures bandwidth, not sustained max RPM. No direct 600 RPM test yet."),
    ("R-12", "Integratable", "Support either shaft or bolted plate outputs",
     "Shafts or bolted plates",
     "Soft", 12, "Demonstration", "Not Verified", "",
     "Design review / demo of integration options."),
    ("R-13", "Torque Density (volume)", "High torque output relative to actuator volume",
     "No specific target set",
     "Soft", 13, "Analysis", "Not Verified", "",
     "Compare to commercial QDDs (e.g., CubeMars AK70-10)."),
]

for r in requirements:
    ws_req.append(r)

# Validation dropdowns
add_validation(ws_req, "E", ["Hard", "Soft"])
add_validation(ws_req, "G", ["Test", "Analysis", "Demonstration", "Inspection"])
add_validation(ws_req, "H", ["Not Verified", "In Progress", "Pass", "Fail", "Partial"])

# Conditional formatting for Status column (H)
from openpyxl.formatting.rule import CellIsRule
ws_req.conditional_formatting.add(
    "H2:H50", CellIsRule(operator="equal", formula=['"Pass"'], fill=PASS_FILL))
ws_req.conditional_formatting.add(
    "H2:H50", CellIsRule(operator="equal", formula=['"Fail"'], fill=FAIL_FILL))
ws_req.conditional_formatting.add(
    "H2:H50", CellIsRule(operator="equal", formula=['"In Progress"'], fill=INPROG_FILL))
ws_req.conditional_formatting.add(
    "H2:H50", CellIsRule(operator="equal", formula=['"Partial"'], fill=PARTIAL_FILL))
ws_req.conditional_formatting.add(
    "H2:H50", CellIsRule(operator="equal", formula=['"Not Verified"'], fill=LIGHT_GRAY))

for row in ws_req.iter_rows(min_row=2, max_row=len(requirements) + 1, max_col=len(req_headers)):
    for cell in row:
        cell.alignment = WRAP
        cell.border = THIN_BORDER

auto_width(ws_req)
ws_req.column_dimensions["C"].width = 45
ws_req.column_dimensions["J"].width = 50


# ═══════════════════════════════════════════════════════════════════════
# SHEET 2: Traceability Matrix (RTM)
# ═══════════════════════════════════════════════════════════════════════
ws_rtm = wb.create_sheet("Traceability Matrix")

# Test info for column headers — grouped by phase
tests_by_phase = [
    # Phase 0
    ("T-009", "Ph0"), ("T-010", "Ph0"), ("T-011", "Ph0"), ("T-014", "Ph0"),
    # Phase 1
    ("T-012", "Ph1"), ("T-013", "Ph1"),
    # Phase 2
    ("T-006", "Ph2"), ("T-007", "Ph2"), ("T-008", "Ph2"),
    # Phase 3
    ("T-015", "Ph3"), ("T-016", "Ph3"),
    # Phase 4
    ("T-017", "Ph4"), ("T-018", "Ph4"), ("T-019", "Ph4"),
    # Rev 00A (completed)
    ("T-001", "00A"), ("T-002", "00A"), ("T-003", "00A"), ("T-004", "00A"), ("T-005", "00A"),
]

test_names = {
    "T-001": "Root Cause Mesh",
    "T-002": "Lid Drag",
    "T-003": "Shoulder Deform",
    "T-004": "Carrier Index",
    "T-005": "Bolt vs Drag",
    "T-006": "Friction Model",
    "T-007": "Step Response",
    "T-008": "Backdrive Torque",
    "T-009": "Motor Kt",
    "T-010": "Motor Cogging",
    "T-011": "Motor Friction",
    "T-012": "Backlash",
    "T-013": "Hand Backdrive",
    "T-014": "Motor Step Resp",
    "T-015": "Plant ID",
    "T-016": "Bode Plot",
    "T-017": "Virtual Spring",
    "T-018": "Spring-Damper",
    "T-019": "Variable Imped",
}

# RTM mapping: (req_id, test_id) -> P/S
# P = primary verification, S = supporting evidence
rtm_map = {
    ("R-01", "T-012"): "P",
    ("R-01", "T-001"): "S",
    ("R-01", "T-004"): "S",
    ("R-04", "T-013"): "P",
    ("R-04", "T-008"): "P",
    ("R-04", "T-006"): "P",
    ("R-04", "T-001"): "S",
    ("R-04", "T-002"): "S",
    ("R-04", "T-005"): "S",
    ("R-05", "T-003"): "S",
    ("R-05", "T-005"): "S",
    ("R-06", "T-003"): "S",
    ("R-06", "T-005"): "S",
    ("R-09", "T-006"): "S",
    ("R-09", "T-011"): "S",
    ("R-11", "T-016"): "S",
}

# Row 1: Phase headers (merged later if desired)
rtm_row1 = ["", ""]  # Req ID, Requirement
for tid, phase in tests_by_phase:
    rtm_row1.append(phase)

# Row 2: Test ID + name headers
rtm_row2 = ["Req ID", "Requirement"]
for tid, _ in tests_by_phase:
    rtm_row2.append(f"{tid}\n{test_names[tid]}")

# Add coverage column
rtm_row1.append("")
rtm_row2.append("Coverage")

ws_rtm.append(rtm_row1)
ws_rtm.append(rtm_row2)

# Style header rows
for col in range(1, len(rtm_row2) + 1):
    for r in [1, 2]:
        cell = ws_rtm.cell(row=r, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

ws_rtm.freeze_panes = "C3"

# Data rows
req_ids = [r[0] for r in requirements]
req_names = [r[1] for r in requirements]

for i, (rid, rname) in enumerate(zip(req_ids, req_names)):
    row_data = [rid, rname]
    has_primary = False
    has_support = False
    for tid, _ in tests_by_phase:
        val = rtm_map.get((rid, tid), "")
        row_data.append(val)
        if val == "P":
            has_primary = True
        if val == "S":
            has_support = True

    # Coverage status
    if has_primary:
        coverage = "Covered"
    elif has_support:
        coverage = "Partial"
    else:
        # Check if it's inspection/analysis (doesn't need a test)
        vmethod = requirements[i][6]
        if vmethod in ("Inspection", "Analysis", "Demonstration"):
            coverage = f"Via {vmethod}"
        else:
            coverage = "Gap"
    row_data.append(coverage)
    ws_rtm.append(row_data)

    row_num = i + 3  # data starts at row 3
    for col_idx in range(1, len(row_data) + 1):
        cell = ws_rtm.cell(row=row_num, column=col_idx)
        cell.alignment = CENTER
        cell.border = THIN_BORDER

        val = cell.value
        if val == "P":
            cell.fill = PRIMARY_FILL
            cell.font = Font(bold=True)
        elif val == "S":
            cell.fill = SUPPORT_FILL
        elif val == "Covered":
            cell.fill = PASS_FILL
        elif val == "Partial":
            cell.fill = PARTIAL_FILL
        elif val == "Gap":
            cell.fill = GAP_FILL
            cell.font = Font(bold=True, color="C00000")
        elif val and val.startswith("Via "):
            cell.fill = LIGHT_BLUE

# Bottom summary row: which requirements each test addresses
summary_row = ["", "Traces To →"]
for col_idx, (tid, _) in enumerate(tests_by_phase):
    traces = [rid for rid in req_ids if rtm_map.get((rid, tid))]
    summary_row.append(", ".join(traces) if traces else "Char.")
summary_row.append("")
ws_rtm.append(summary_row)
summary_row_num = len(req_ids) + 3
for col in range(1, len(summary_row) + 1):
    cell = ws_rtm.cell(row=summary_row_num, column=col)
    cell.font = Font(bold=True, italic=True, size=9)
    cell.fill = LIGHT_GRAY
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# Column widths
ws_rtm.column_dimensions["A"].width = 8
ws_rtm.column_dimensions["B"].width = 18
for col in range(3, len(rtm_row2)):
    ws_rtm.column_dimensions[get_column_letter(col)].width = 13
ws_rtm.column_dimensions[get_column_letter(len(rtm_row2))].width = 14


# ═══════════════════════════════════════════════════════════════════════
# SHEET 3: Test Log
# ═══════════════════════════════════════════════════════════════════════
ws_tests = wb.create_sheet("Test Log")
test_headers = [
    "Test ID", "Test Name", "Phase", "Traces To", "Procedure Reference",
    "Prerequisites", "Status", "Date Executed", "Result",
    "Acceptance Criteria", "Pass/Fail", "Notes / Actions",
]
ws_tests.append(test_headers)
style_header(ws_tests, len(test_headers))

test_data = [
    ("T-001", "Root cause analysis — tight gear meshing", "Rev 00A", "R-01, R-04",
     "test-tracker.md T-001", "Part dimensions",
     "Complete", "2026-03-13", "Herringbone alignment error identified",
     "Dimensions within printer tolerance", "Partial",
     "Root cause: herringbone misalignment. Action: spur gears for Rev 00B. Dimensional measurements still pending."),
    ("T-002", "Lid drag — over-constraint on carrier", "Rev 00A", "R-04, R-09",
     "test-tracker.md T-002", "None",
     "Complete", "2026-03-13", "Root cause identified — tolerance stackup",
     "Lid does not add parasitic drag", "Fail",
     "Lid double-duty causes stackup. Design change: decouple ring gear clamping from carrier constraint."),
    ("T-003", "Carrier bearing shoulder deformation", "Rev 00A", "R-05, R-06",
     "test-tracker.md T-003", "None",
     "Complete", "2026-03-13", "Print quality issue on top carrier shoulders",
     "Shoulder change < 0.1 mm, no cracking", "Inconclusive",
     "Bottom shoulders good. Top shoulders have print artifacts (orientation trade-off). No dimensional measurement yet."),
    ("T-004", "Carrier half indexing — clocking under load", "Rev 00A", "R-01",
     "test-tracker.md T-004", "T-003",
     "Complete", "2026-03-13", "~0.5 mm play exists but no operational driving force",
     "No visible relative motion under hand torsion", "Pass",
     "Load path analysis: both halves driven by same pins. Clocking not a realistic failure mode."),
    ("T-005", "Carrier bolt torque vs planet resistance", "Rev 00A", "R-04, R-05",
     "test-tracker.md T-005", "None",
     "Complete", "2026-03-13", "Narrow window between carrier play and planet binding",
     "Planets spin freely across bolt torque range", "Fail",
     "Shoulders sensitive to clamping force. 1.1 Nm → drag. Backed off → free. Rev 00B shoulder redesign needed."),
    ("T-006", "Friction characterization — Coulomb + viscous", "Phase 2", "R-04, R-09",
     "test-campaign-rev00b.md Phase 2", "Motor + encoder + ODrive, gearbox assembled",
     "Not Started", "", "",
     "Reflected friction torque < 1 Nm; linear friction model", "",
     "Motor torque current at constant velocities. Subtract T-011 baseline for gearbox-only friction."),
    ("T-007", "Step response — system dynamics", "Phase 2", "—",
     "test-campaign-rev00b.md Phase 2", "Motor + encoder + ODrive, gearbox assembled",
     "Not Started", "", "",
     "System settles, overshoot < 10%, SSE < 0.5°", "",
     "Characterization test. Compare to T-014 motor-only baseline."),
    ("T-008", "Backdriving torque — passive resistance", "Phase 2", "R-04",
     "test-campaign-rev00b.md Phase 2", "Motor + encoder + ODrive, gearbox assembled",
     "Not Started", "", "",
     "Backdriving torque < 1 Nm at output", "",
     "Quantitative version of T-013. Motor in idle, turn output by hand."),
    ("T-009", "Motor Kt verification (baseline)", "Phase 0", "—",
     "test-campaign-rev00b.md Phase 0", "ODrive + motor + encoder configured",
     "Not Started", "", "",
     "No-load Iq < 1A at 2 rev/s", "",
     "Baseline sanity check. Every downstream torque calc depends on Kt."),
    ("T-010", "Motor cogging & friction profile", "Phase 0", "—",
     "test-campaign-rev00b.md Phase 0", "T-009 pass",
     "Not Started", "", "",
     "Periodic cogging, roughly symmetric CW/CCW", "",
     "Motor-only baseline. Needed to isolate gearbox effects in T-006."),
    ("T-011", "Motor friction vs speed — no load", "Phase 0", "—",
     "test-campaign-rev00b.md Phase 0", "T-009 pass",
     "Not Started", "", "",
     "Roughly linear friction vs speed", "",
     "Motor-only Coulomb + viscous model. Baseline for T-006 subtraction."),
    ("T-012", "Output shaft backlash measurement", "Phase 1", "R-01",
     "test-campaign-rev00b.md Phase 1", "Gearbox assembled, motor shaft locked",
     "Not Started", "", "",
     "Backlash ≤ 0.5° (R-01 hard req)", "",
     "Dial indicator + lever arm at 4+ positions around one revolution."),
    ("T-013", "Hand backdriving — qualitative + rough torque", "Phase 1", "R-04",
     "test-campaign-rev00b.md Phase 1", "Gearbox assembled",
     "Not Started", "", "",
     "Backdriving torque < 1 Nm (R-04)", "",
     "Kitchen scale + 200 mm lever arm. Quick check before powered T-008."),
    ("T-014", "Motor step response (baseline)", "Phase 0", "—",
     "test-campaign-rev00b.md Phase 0", "T-009 pass",
     "Not Started", "", "",
     "Clean second-order response, consistent across 3–5 reps", "",
     "Motor-only baseline for T-007 comparison. Extract tr, OS%, ts, SSE."),
    ("T-015", "Plant identification from step response data", "Phase 3", "—",
     "test-campaign-rev00b.md Phase 3", "T-014, T-007 complete",
     "Not Started", "", "",
     "Model matches measured step response within 10%", "",
     "Fit G(s) = 1/(Js² + bs). Use torque mode for open-loop plant ID."),
    ("T-016", "Frequency response — Bode plot", "Phase 3", "—",
     "test-campaign-rev00b.md Phase 3", "T-015 (for validation)",
     "Not Started", "", "",
     "Measured Bode matches predicted from T-015", "",
     "Chirp or discrete sine dwell. Motor alone AND motor+gearbox overlay."),
    ("T-017", "Impedance control — virtual spring", "Phase 4", "—",
     "test-campaign-rev00b.md Phase 4", "T-015 (plant model for tuning)",
     "Not Started", "", "",
     "Linear force-displacement, smooth feel, returns to setpoint", "",
     "Tier A: use ODrive internal PID as spring-damper. Start here."),
    ("T-018", "Impedance control — spring-damper", "Phase 4", "—",
     "test-campaign-rev00b.md Phase 4", "T-017",
     "Not Started", "", "",
     "Damped return to setpoint, higher K achievable with B", "",
     "Tier B: Python torque loop (~200 Hz). True impedance control."),
    ("T-019", "Variable impedance demo (stretch goal)", "Phase 4", "—",
     "test-campaign-rev00b.md Phase 4", "T-018",
     "Not Started", "", "",
     "Smooth real-time parameter adjustment, no instability", "",
     "Tkinter GUI with K/B sliders. Portfolio hero content."),
]

for t in test_data:
    ws_tests.append(t)

add_validation(ws_tests, "G", ["Not Started", "In Progress", "Complete", "Blocked"])
add_validation(ws_tests, "K", ["Pass", "Fail", "Inconclusive", "Partial", ""])

# Conditional formatting for Status (G) and Pass/Fail (K)
ws_tests.conditional_formatting.add(
    "G2:G50", CellIsRule(operator="equal", formula=['"Complete"'], fill=PASS_FILL))
ws_tests.conditional_formatting.add(
    "G2:G50", CellIsRule(operator="equal", formula=['"In Progress"'], fill=INPROG_FILL))
ws_tests.conditional_formatting.add(
    "G2:G50", CellIsRule(operator="equal", formula=['"Blocked"'], fill=FAIL_FILL))
ws_tests.conditional_formatting.add(
    "G2:G50", CellIsRule(operator="equal", formula=['"Not Started"'], fill=LIGHT_GRAY))

ws_tests.conditional_formatting.add(
    "K2:K50", CellIsRule(operator="equal", formula=['"Pass"'], fill=PASS_FILL))
ws_tests.conditional_formatting.add(
    "K2:K50", CellIsRule(operator="equal", formula=['"Fail"'], fill=FAIL_FILL))
ws_tests.conditional_formatting.add(
    "K2:K50", CellIsRule(operator="equal", formula=['"Inconclusive"'], fill=INPROG_FILL))

for row in ws_tests.iter_rows(min_row=2, max_row=len(test_data) + 1, max_col=len(test_headers)):
    for cell in row:
        cell.alignment = WRAP
        cell.border = THIN_BORDER

auto_width(ws_tests)
ws_tests.column_dimensions["B"].width = 35
ws_tests.column_dimensions["E"].width = 30
ws_tests.column_dimensions["F"].width = 30
ws_tests.column_dimensions["I"].width = 35
ws_tests.column_dimensions["J"].width = 35
ws_tests.column_dimensions["L"].width = 50


# ═══════════════════════════════════════════════════════════════════════
# SHEET 4: Unknowns Register
# ═══════════════════════════════════════════════════════════════════════
ws_unk = wb.create_sheet("Unknowns Register")
unk_headers = [
    "Unknown ID", "Description", "Category", "Priority",
    "Status", "Linked Tests", "Resolution", "Date Resolved",
]
ws_unk.append(unk_headers)
style_header(ws_unk, len(unk_headers))

unknowns = [
    ("U-01", "Planet positioning & mesh clearance — geometry vs. assembly?",
     "Assembly/fit", "Test first",
     "Postponed → Rev 00B", "T-001", "Herringbone misalignment identified as root cause. Switched to spur gears.", ""),
    ("U-02", "(Merged into U-01)", "—", "—", "Merged", "", "", ""),
    ("U-03", "Lid causing drag — over-constraint on carrier?",
     "Assembly/fit", "Test first",
     "Root cause identified", "T-002",
     "Tolerance stackup across carrier assembly causes lid to press on carrier top. Redesign for Rev 00B.", "2026-03-13"),
    ("U-04", "Carrier bearing shoulder deformation under bolt torque",
     "Structural", "Test first",
     "Confirmed", "T-003, T-005",
     "Shoulders sensitive to torque. Print quality issue on top carrier. Rev 00B redesign.", "2026-03-13"),
    ("U-05", "Carrier halves need indexing?",
     "Assembly/fit", "Monitor/skip",
     "Resolved", "T-004",
     "No operational clocking load — both halves driven by same pins.", "2026-03-13"),
    ("U-06", "Gearbox structural capacity (handles motor torque × 5:1?)",
     "Structural", "After fundamentals",
     "Not Started", "", "", ""),
    ("U-07", "Output shaft strength — self-tap holding, diameter adequate",
     "Structural", "Verify early",
     "Known Issue", "",
     "Known undersized → enlarged in Rev 00B design.", ""),
    ("U-08", "Backlash — quantified",
     "Performance", "After fundamentals",
     "Not Started", "T-012",
     "", ""),
    ("U-09", "Backdrivability — quantified",
     "Performance", "After fundamentals",
     "Not Started", "T-013, T-006, T-008",
     "", ""),
    ("U-10", "Durability under sustained use",
     "Degradation", "After fundamentals",
     "Not Started", "",
     "", ""),
]

for u in unknowns:
    ws_unk.append(u)

add_validation(ws_unk, "C", ["Assembly/fit", "Structural", "Performance", "Degradation", "—"])
add_validation(ws_unk, "D", ["Test first", "After fundamentals", "Verify early", "Monitor/skip", "—"])
add_validation(ws_unk, "E", ["Not Started", "In Progress", "Resolved", "Postponed",
                              "Confirmed", "Root cause identified", "Known Issue", "Merged"])

ws_unk.conditional_formatting.add(
    "E2:E50", CellIsRule(operator="equal", formula=['"Resolved"'], fill=PASS_FILL))
ws_unk.conditional_formatting.add(
    "E2:E50", CellIsRule(operator="equal", formula=['"Not Started"'], fill=LIGHT_GRAY))
ws_unk.conditional_formatting.add(
    "E2:E50", CellIsRule(operator="equal", formula=['"Confirmed"'], fill=FAIL_FILL))
ws_unk.conditional_formatting.add(
    "E2:E50", CellIsRule(operator="equal", formula=['"Root cause identified"'], fill=INPROG_FILL))

for row in ws_unk.iter_rows(min_row=2, max_row=len(unknowns) + 1, max_col=len(unk_headers)):
    for cell in row:
        cell.alignment = WRAP
        cell.border = THIN_BORDER

auto_width(ws_unk)
ws_unk.column_dimensions["B"].width = 50
ws_unk.column_dimensions["G"].width = 55


# ═══════════════════════════════════════════════════════════════════════
# SHEET 5: Health Log
# ═══════════════════════════════════════════════════════════════════════
ws_health = wb.create_sheet("Health Log")
health_headers = [
    "Date", "Session #", "Backlash (°)", "Friction Torque (Nm)",
    "Ambient Temp (°C)", "Visual Condition", "Cumulative Runtime (min)", "Notes",
]
ws_health.append(health_headers)
style_header(ws_health, len(health_headers))

# Add a few placeholder rows to show format
ws_health.append(["", 1, "", "", "", "", 0, "Initial baseline — record before first powered test"])

for row in ws_health.iter_rows(min_row=2, max_row=2, max_col=len(health_headers)):
    for cell in row:
        cell.alignment = WRAP
        cell.border = THIN_BORDER
        cell.font = Font(italic=True, color="808080")

auto_width(ws_health)
ws_health.column_dimensions["F"].width = 35
ws_health.column_dimensions["H"].width = 40


# ═══════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════
output_path = r"C:\Users\aaron\Documents\c-projects\qdd-gearbox\testing\qdd-rtm.xlsx"
wb.save(output_path)
print(f"Created: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"Requirements: {len(requirements)}")
print(f"Tests: {len(test_data)}")
print(f"Unknowns: {len(unknowns)}")
